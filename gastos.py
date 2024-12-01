from openpyxl import Workbook, load_workbook

# Carrega ou cria o arquivo Excel
try:
    wb = load_workbook("gastos.xlsx")
    sheet = wb.active
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Gastos"

# Pergunta quantas pessoas estão na divisão
quantidadeDePessoas = int(input("Quantas pessoas estão na divisão? "))

# Adiciona os nomes das pessoas na primeira linha (cabeçalho)
for coluna in range(1, quantidadeDePessoas + 1):
    nome = input(f"Digite o nome da pessoa {coluna}: ")
    sheet.cell(row=1, column=coluna).value = nome

# Adiciona os gastos para cada pessoa
continuar = "s"
while continuar.lower() == "s":
    print("\nAdicione um gasto:")
    
    # Escolha a pessoa
    for coluna in range(1, quantidadeDePessoas + 1):
        nome = sheet.cell(row=1, column=coluna).value
        print(f"{coluna}. {nome}")
    
    escolha = int(input("Digite o número da pessoa: "))
    
    # Verifica se a escolha é válida
    if escolha < 1 or escolha > quantidadeDePessoas:
        print("Escolha inválida. Tente novamente.")
        continue
    
    # Adiciona o gasto
    pessoa_coluna = escolha
    valor = float(input("Digite o valor do gasto: "))
    
    # Encontra a próxima linha vazia na coluna da pessoa
    linha = 2
    while sheet.cell(row=linha, column=pessoa_coluna).value is not None:
        linha += 1
    
    # Adiciona o gasto
    sheet.cell(row=linha, column=pessoa_coluna).value = valor
    print(f"Gasto de R${valor:.2f} adicionado para {sheet.cell(row=1, column=pessoa_coluna).value}.")

    # Pergunta se deseja continuar adicionando gastos
    continuar = input("\nDeseja adicionar outro gasto? (s/n): ")

# Calcula o total de gastos por pessoa
linha_total = sheet.max_row + 1
for coluna in range(1, quantidadeDePessoas + 1):
    total = sum(sheet.cell(row=row, column=coluna).value or 0 for row in range(2, linha_total))
    sheet.cell(row=linha_total, column=coluna).value = f"Total: R${total:.2f}"

# Salva as alterações no arquivo
wb.save("gastos.xlsx")
print("\nGastos registrados e arquivo salvo como 'gastos.xlsx'.")
